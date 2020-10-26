#!/usr/bin/env python
# coding: utf-8

# In[2]:


from datetime import datetime, timedelta
import os
import openpyxl
import os
import pandas as pd

ROOT_DIR = './'

MARKS = ['СКС', "СКД", "ДСТ", "СБС","ДССК"]
print("Ready!")


# In[3]:


#Загрузка образца отчёта
wb = openpyxl.load_workbook('weekly_basis.xlsx')
report_sheet = wb['Форма горизонталь1']


report_titles = report_sheet['A1':'AL3']
hiden_data = report_sheet['A4':'G8']
rows = {'sks':report_sheet['G4':'AL4'],'skd':report_sheet['G5':'AL5'],'tep35':report_sheet['G6':'AL6'],
        'tep50':report_sheet['G7':'AL7'],'total':report_sheet['G8':'AL8']}


# In[5]:


#Подгрузка отчётов из SAP 

#добавить проверку на верность даты начала и конца выгрузки
prod_report  = pd.read_excel(os.path.join(ROOT_DIR, 'ZPP_PROD_REPORT.XLSX'),sheet_name = 'Sheet1', thousands = ',').drop(columns='Продукт').rename(columns={'Продукт.1':'Продукт'})
mb52_report  = pd.read_excel(os.path.join(ROOT_DIR, 'MB52.XLSX'),sheet_name = 'Sheet1', thousands = ',')


# In[6]:


#Используемые функции(отлаженные и протестированные)
#печатает слайс
def print_row(row_slice):
    for cell_obj in row_slice:
        for cell in cell_obj:
            print(cell.value, cell.column, cell.coordinate)
#создает эксель, в который пишутся результаты
def create_resulting_wb():
    wb_result = openpyxl.Workbook()
    wb_result.create_sheet(title = 'Отчёт', index = 0)
    result_sheet = wb_result['Отчёт']
    for irow in range(8):
        result_sheet.append([])
    wb_result.save('weekly_report.xlsx')
    return wb_result
#копирую инфу из одного экселя в другой
def copy_cells(slice_to_copy,sheet_we_copy_in):
    for cells_obj in slice_to_copy:
        for cell in cells_obj:
            cell_tmp = sheet_we_copy_in.cell(row = cell.row, column = cell.column)
            cell_tmp.value = cell.value
    wb_result.save('weekly_report.xlsx')
#Посчитать последнюю строку    
def find_total():
    total_positions = []
    for cell_obj in rows['total']:
        for cell in cell_obj:
            if type(cell.value) == str:
                total_positions.append((cell.value, cell.column, cell.coordinate))
    for i in total_positions:
        try:
            report_sheet[i[2]] = (
                                  float(report_sheet.cell(column=i[1],row=4).value) + 
                                  float(report_sheet.cell(column=i[1],row=5).value) + 
                                  float(report_sheet.cell(column=i[1],row=6).value) + 
                                  float(report_sheet.cell(column=i[1],row=7).value)
                                  )
        except ValueError:
            #print(i[2])
            report_sheet[i[2]] = ''
    #return report_sheet['G8':'AL8']

def put_info_in_col(col_n, data_dict, sheet_we_put_in):
    cell_tmp = sheet_we_put_in.cell(row = 4, column = col_n)#CKC
    cell_tmp.value =data_dict['СКС']
    cell_tmp = sheet_we_put_in.cell(row = 5, column = col_n)#CKC
    cell_tmp.value =data_dict['СКД']
    cell_tmp = sheet_we_put_in.cell(row = 6, column = col_n)#CKC
    cell_tmp.value =data_dict['ДСТ']
    cell_tmp = sheet_we_put_in.cell(row = 7, column = col_n)#CKC
    cell_tmp.value =data_dict['СБС']
    wb_result.save('weekly_report.xlsx')


#Получаем данные для 17стобца из ZPP_PROD_REPORT -> вариант VSK_PERIOD_PRD (Отчёт ВСК производства и сортировки продукции. Паспортизация)
def get_column_17_info(df_zpp_prod_report):
    #['СКС', "СКД"+"ДССК", "ДСТ", "СБС"]
    RESULT = {'СКС':0,'СКД':0,'ДСТ':0,'СБС':0}
    for i in range(prod_report.shape[0]):
        row_tmp = prod_report.iloc[i,:]
        row_tmp['Продукт'] = str(row_tmp['Продукт'])
        if MARKS[0] in row_tmp['Продукт'] and not 'ЭП' in row_tmp['Продукт']:#CKC - верно
            RESULT['СКС'] += row_tmp['Паспортизация'].astype(float) 
            continue
        if MARKS[1] in row_tmp['Продукт'] and not 'ЭП' in row_tmp['Продукт']:#СКД - верно
            RESULT['СКД'] += row_tmp['Паспортизация'].astype(float)
            continue
        if MARKS[4] in row_tmp['Продукт'] and not 'ЭП' in row_tmp['Продукт']:#ДССК - верно
            RESULT['СКД'] += row_tmp['Паспортизация'].astype(float)
            continue
        if MARKS[2] in row_tmp['Продукт']:# and not 'ЭП' in row_tmp['Продукт']:#ДСТ
            RESULT['ДСТ'] += row_tmp['Паспортизация'].astype(float)
            continue
        if MARKS[3] in row_tmp['Продукт']:# and not 'ЭП' in row_tmp['Продукт']:#СБС
            RESULT['СБС'] += row_tmp['Паспортизация'].astype(float) 
            continue
    return RESULT
#Получаем данные для 10стобца из MB52 -> вариант BOGDANOVA


# In[62]:


#Основной список вызов функций (main)
wb_result = create_resulting_wb()
result_sheet = wb_result['Отчёт']
copy_cells(report_titles, result_sheet)
copy_cells(hiden_data, result_sheet)
#17 столбец
put_info_in_col(17,get_column_17_info(prod_report),result_sheet)
#последняя строка
find_total()
copy_cells(rows['total'], result_sheet)


# In[54]:


#MARKS = ['СКС', "СКД", "ДСТ", "СБС","ДССК"]


# In[24]:


mb52_report[(mb52_report['Краткий текст материала'].str.contains(i) == True) & (mb52_report['Краткий текст материала'].str.contains('ЭП') == False) & (mb52_report['Краткий текст материала'].str.contains(' 611') == False)]


# In[30]:


#rail = ORION[(ORION['Метод отгрузки'] == 'ЖД -вагон') & (ORION['Тип объекта'] == 'РО')][[date,'Наименование   позиции']].dropna().reset_index().drop(columns = 'index')
def get_column10_info(df_mb52_report):
    RESULT = {'СКС':0,'СКД':0,'ДСТ':0,'СБС':0}#готовый продукт без ЭП и брака 611
    print(RESULT['СКД'])
    for i in MARKS:
    #RESULT['СКС'] = df_mb52_report[(df_mb52_report['Краткий текст материала'].str.contains('СКС') == True) & (df_mb52_report['Краткий текст материала'].str.contains('ЭП') == False) & (df_mb52_report['Краткий текст материала'].str.contains(' 611') == False)]['СвобИспользЗапас'].sum()#.reset_index().drop(columns = 'index')
        if i == 'ДССК':
            RESULT['СКД'] += df_mb52_report[(df_mb52_report['Краткий текст материала'].str.contains(i) == True) & (df_mb52_report['Краткий текст материала'].str.contains(' 611') == False)]['СвобИспользЗапас'].sum()#.reset_index().drop(columns = 'index')
        else:
            if(i == 'СБС') or (i =='ДСТ'):
                RESULT[i] += df_mb52_report[(df_mb52_report['Краткий текст материала'].str.contains(i) == True)  & (df_mb52_report['Краткий текст материала'].str.contains(' 611') == False)]['СвобИспользЗапас'].sum()#.reset_index().drop(columns = 'index')
            else:
                RESULT[i] += df_mb52_report[(df_mb52_report['Краткий текст материала'].str.contains(i) == True) & (df_mb52_report['Краткий текст материала'].str.contains(' 611') == False)]['СвобИспользЗапас'].sum()
    return RESULT
    
get_column10_info(mb52_report)    
    
    
    
    
    
    


# In[30]:


RESULT = [0,0]
for i in range(prod_report.shape[0]):
    row_tmp = prod_report.iloc[i,:]
    row_tmp['Продукт'] = str(row_tmp['Продукт'])
    if MARKS[1] in row_tmp['Продукт'] and not 'ЭП' in row_tmp['Продукт']:#СКД
        RESULT[1] += row_tmp['Паспортизация'].astype(float)
        print(row_tmp)


# In[45]:


RESULT1 = [0,0,0]
for i in range(prod_report.shape[0]):
    row_tmp = prod_report.iloc[i,:]
    row_tmp['Продукт'] = str(row_tmp['Продукт'])
    if z in row_tmp['Продукт'] and not 'ЭП' in row_tmp['Продукт']:#ДССК
        RESULT1[2] += row_tmp['Паспортизация'].astype(float)
        print(row_tmp)        
RESULT1


# In[44]:


z = prod_report.iloc[1370,0].split(' ')[1]
z


# In[ ]:




